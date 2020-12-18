from selenium import webdriver
from openpyxl import *

# QUALI ONE

url = 'https://fiaresultsandstatistics.motorsportstats.com/results/2020-abu-dhabi-grand-prix/classification/68a407e5-2fd4-4797-afbb-5159450944e2'
driver = webdriver.Chrome('/Users/siddhantsingh/Documents/Onedrive/CSProjects/PythonProjects/chromedriver')
driver.get(url)
racers = driver.find_elements_by_class_name("_2xhp6")

HAM = []
BOT = []
VET = []
LEC = []
VER = []
ALB = []
NOR = []
SAI = []
RIC = []
OCO = []
GAS = []
KVY = []
MAG = []
FIT = []
RUS = []
LAT = []
RAI = []
GIO = []
PER = []
STR = []

for racer in racers:
    print(racer.text.split(" "))
    if racer.text.split(" ")[1] == '44':
        HAM.append(str(racer.text.split(" ")[12]))
    elif racer.text.split(" ")[1] == '77':
        BOT.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '33':
        VER.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '23':
        ALB.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '4':
        NOR.append(str(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '11':
        PER.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '18':
        STR.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '3':
        RIC.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '55':
        SAI.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '26':
        KVY.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '7':
        RAI.append(str(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '20':
        MAG.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '51':
        FIT.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '99':
        GIO.append(str(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '10':
        GAS.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '63':
        RUS.append(str(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '6':
        LAT.append(str(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '31':
        OCO.append(str(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '16':
        LEC.append(str(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '5':
        VET.append(str(racer.text.split(" ")[8]))


file_path = "/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx"
wb = load_workbook(file_path)
s = wb['Q17']

colx = 2
rowx = 3


for rowx in range(3, 23):
    timeCol = 7
    if s.cell(rowx, colx).value == 'HAM':
        if len(HAM) != 0:
            time = HAM[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'BOT':
        if len(BOT) != 0:
            time = BOT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VET':
        if len(VET) != 0:
            time = VET[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LEC':
        if len(LEC) != 0:
            time = LEC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VER':
        if len(VER) != 0:
            time = VER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'ALB':
        if len(ALB) != 0:
            time = ALB[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'NOR':
        if len(NOR) != 0:
            time = NOR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'SAI':
        if len(SAI) != 0:
            time = SAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RIC':
        if len(RIC) != 0:
            time = RIC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'OCO':
        if len(OCO) != 0:
            time = OCO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GAS':
        if len(GAS) != 0:
            time = GAS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'KVY':
        if len(KVY) != 0:
            time = KVY[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'MAG':
        if len(MAG) != 0:
            time = MAG[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'FIT':
        if len(FIT) != 0:
            time = FIT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RUS':
        if len(RUS) != 0:
            time = RUS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LAT':
        if len(LAT) != 0:
            time = LAT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GIO':
        if len(GIO) != 0:
            time = GIO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RAI':
        if len(RAI) != 0:
            time = RAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'PER':
        if len(PER) != 0:
            time = PER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'STR':
        if len(STR) != 0:
            time = STR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'


wb.save('/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx')
print("\n")

# QUALI TWO

url = 'https://fiaresultsandstatistics.motorsportstats.com/results/2020-abu-dhabi-grand-prix/classification/a243e285-0851-4aa1-a0e5-55fcc8c394e5'
driver = webdriver.Chrome('/Users/siddhantsingh/Documents/Onedrive/CSProjects/PythonProjects/chromedriver')
driver.get(url)
racers = driver.find_elements_by_class_name("_2xhp6")

HAM = []
BOT = []
VET = []
LEC = []
VER = []
ALB = []
NOR = []
SAI = []
RIC = []
OCO = []
GAS = []
KVY = []
MAG = []
FIT = []
RUS = []
LAT = []
RAI = []
GIO = []
PER = []
STR = []

for racer in range(15):
    print(racers[racer].text.split(" "))
    if racers[racer].text.split(" ")[1] == '44':
        HAM.append(str(racers[racer].text.split(" ")[12]))
    elif racers[racer].text.split(" ")[1] == '77':
        BOT.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '33':
        VER.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '23':
        ALB.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '4':
        NOR.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '11':
        PER.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '18':
        STR.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '3':
        RIC.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '55':
        SAI.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '26':
        KVY.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '7':
        RAI.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '20':
        MAG.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '51':
        FIT.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '99':
        GIO.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '10':
        GAS.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '63':
        RUS.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '6':
        LAT.append(str(racers[racer].text.split(" ")[8]))
    elif racers[racer].text.split(" ")[1] == '31':
        OCO.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '16':
        LEC.append(str(racers[racer].text.split(" ")[8]))
    elif racers[racer].text.split(" ")[1] == '5':
        VET.append(str(racers[racer].text.split(" ")[8]))

file_path = "/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx"
wb = load_workbook(file_path)
s = wb['Q17']

colx = 2
rowx = 3

for rowx in range(3, 23):
    timeCol = 8
    if s.cell(rowx, colx).value == 'HAM':
        if len(HAM) != 0:
            time = HAM[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'BOT':
        if len(BOT) != 0:
            time = BOT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VET':
        if len(VET) != 0:
            time = VET[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LEC':
        if len(LEC) != 0:
            time = LEC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VER':
        if len(VER) != 0:
            time = VER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'ALB':
        if len(ALB) != 0:
            time = ALB[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'NOR':
        if len(NOR) != 0:
            time = NOR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'SAI':
        if len(SAI) != 0:
            time = SAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RIC':
        if len(RIC) != 0:
            time = RIC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'OCO':
        if len(OCO) != 0:
            time = OCO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GAS':
        if len(GAS) != 0:
            time = GAS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'KVY':
        if len(KVY) != 0:
            time = KVY[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'MAG':
        if len(MAG) != 0:
            time = MAG[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'FIT':
        if len(FIT) != 0:
            time = FIT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RUS':
        if len(RUS) != 0:
            time = RUS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LAT':
        if len(LAT) != 0:
            time = LAT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GIO':
        if len(GIO) != 0:
            time = GIO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RAI':
        if len(RAI) != 0:
            time = RAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'PER':
        if len(PER) != 0:
            time = PER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'STR':
        if len(STR) != 0:
            time = STR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

wb.save('/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx')
print("\n")


# QUALI THREE

url = 'https://fiaresultsandstatistics.motorsportstats.com/results/2020-abu-dhabi-grand-prix/classification/ee442193-13a3-40f5-9d1f-41fc6a3208a9'
driver = webdriver.Chrome('/Users/siddhantsingh/Documents/Onedrive/CSProjects/PythonProjects/chromedriver')
driver.get(url)
racers = driver.find_elements_by_class_name("_2xhp6")

HAM = []
BOT = []
VET = []
LEC = []
VER = []
ALB = []
NOR = []
SAI = []
RIC = []
OCO = []
GAS = []
KVY = []
MAG = []
FIT = []
RUS = []
LAT = []
RAI = []
GIO = []
PER = []
STR = []

for racer in range(10):
    print(racers[racer].text.split(" "))
    if racers[racer].text.split(" ")[1] == '44':
        HAM.append(str(racers[racer].text.split(" ")[12]))
    elif racers[racer].text.split(" ")[1] == '77':
        BOT.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '33':
        VER.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '23':
        ALB.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '4':
        NOR.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '11':
        PER.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '18':
        STR.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '3':
        RIC.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '55':
        SAI.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '26':
        KVY.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '7':
        RAI.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '20':
        MAG.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '51':
        FIT.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '99':
        GIO.append(str(racers[racer].text.split(" ")[10]))
    elif racers[racer].text.split(" ")[1] == '10':
        GAS.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '63':
        RUS.append(str(racers[racer].text.split(" ")[9]))
    elif racers[racer].text.split(" ")[1] == '6':
        LAT.append(str(racers[racer].text.split(" ")[8]))
    elif racers[racer].text.split(" ")[1] == '31':
        OCO.append(str(racers[racer].text.split(" ")[11]))
    elif racers[racer].text.split(" ")[1] == '16':
        LEC.append(str(racers[racer].text.split(" ")[8]))
    elif racers[racer].text.split(" ")[1] == '5':
        VET.append(str(racers[racer].text.split(" ")[8]))

file_path = "/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx"
wb = load_workbook(file_path)
s = wb['Q17']

colx = 2
rowx = 3

for rowx in range(3, 23):
    timeCol = 9
    if s.cell(rowx, colx).value == 'HAM':
        if len(HAM) != 0:
            time = HAM[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'BOT':
        if len(BOT) != 0:
            time = BOT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VET':
        if len(VET) != 0:
            time = VET[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LEC':
        if len(LEC) != 0:
            time = LEC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'VER':
        if len(VER) != 0:
            time = VER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'ALB':
        if len(ALB) != 0:
            time = ALB[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'NOR':
        if len(NOR) != 0:
            time = NOR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'SAI':
        if len(SAI) != 0:
            time = SAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RIC':
        if len(RIC) != 0:
            time = RIC[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'OCO':
        if len(OCO) != 0:
            time = OCO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GAS':
        if len(GAS) != 0:
            time = GAS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'KVY':
        if len(KVY) != 0:
            time = KVY[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'MAG':
        if len(MAG) != 0:
            time = MAG[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'FIT':
        if len(FIT) != 0:
            time = FIT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RUS':
        if len(RUS) != 0:
            time = RUS[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'LAT':
        if len(LAT) != 0:
            time = LAT[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'GIO':
        if len(GIO) != 0:
            time = GIO[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'RAI':
        if len(RAI) != 0:
            time = RAI[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'PER':
        if len(PER) != 0:
            time = PER[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

    elif s.cell(rowx, colx).value == 'STR':
        if len(STR) != 0:
            time = STR[0]
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = time
        else:
            timeLoc = s.cell(rowx, timeCol)
            timeLoc.value = 'NA'

wb.save('/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx')
print("\n")
