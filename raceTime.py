from selenium import webdriver
from openpyxl import *

url = 'https://fiaresultsandstatistics.motorsportstats.com/results/2020-bahrain-grand-prix/classification/e05a44dc-d24d-4c21-8912-894b012f6456'
driver = webdriver.Chrome('/Users/siddhantsingh/Documents/Onedrive/CSProjects/PythonProjects/chromedriver')
driver.get(url)
racers = driver.find_elements_by_class_name("_2xhp6")

HAM = []
BOT = []
VET = []
LEC = []
VER = []
ALB = []
NOR = []
SAI = []
RIC = []
OCO = []
GAS = []
KVY = []
MAG = []
# GRO = []
RUS = []
LAT = []
RAI = []
GIO = []
PER = []
STR = []

for racer in racers:
    print(racer.text.split(" "))
    if racer.text.split(" ")[1] == '44':
        if racer.text.split(" ")[0] != 'DNF':
            HAM.append(int(racer.text.split(" ")[0]))
        else:
            HAM.append('DNF')
        HAM.append(str(racer.text.split(" ")[12]))
        HAM.append(int(racer.text.split(" ")[11]))
    elif racer.text.split(" ")[1] == '77':
        if racer.text.split(" ")[0] != 'DNF':
            BOT.append(int(racer.text.split(" ")[0]))
        else:
            BOT.append('DNF')
        BOT.append(str(racer.text.split(" ")[11]))
        BOT.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '33':
        if racer.text.split(" ")[0] != 'DNF':
            VER.append(int(racer.text.split(" ")[0]))
        else:
            VER.append('DNF')
        VER.append(str(racer.text.split(" ")[11]))
        VER.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '23':
        if racer.text.split(" ")[0] != 'DNF':
            ALB.append(int(racer.text.split(" ")[0]))
        else:
            ALB.append('DNF')
        ALB.append(str(racer.text.split(" ")[11]))
        ALB.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '4':
        if racer.text.split(" ")[0] != 'DNF':
            NOR.append(int(racer.text.split(" ")[0]))
        else:
            NOR.append('DNF')
        NOR.append(str(racer.text.split(" ")[10]))
        NOR.append(int(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '11':
        if racer.text.split(" ")[0] != 'DNF':
            PER.append(int(racer.text.split(" ")[0]))
        else:
            PER.append('DNF')
        PER.append(str(racer.text.split(" ")[11]))
        PER.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '18':
        if racer.text.split(" ")[0] != 'DNF':
            STR.append(int(racer.text.split(" ")[0]))
        else:
            STR.append('DNF')
        STR.append(str(racer.text.split(" ")[11]))
        STR.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '3':
        if racer.text.split(" ")[0] != 'DNF':
            RIC.append(int(racer.text.split(" ")[0]))
        else:
            RIC.append('DNF')
        RIC.append(str(racer.text.split(" ")[11]))
        RIC.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '55':
        if racer.text.split(" ")[0] != 'DNF':
            SAI.append(int(racer.text.split(" ")[0]))
        else:
            SAI.append('DNS')
        SAI.append(str(racer.text.split(" ")[9]))
        SAI.append(int(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '26':
        if racer.text.split(" ")[0] != 'DNF':
            KVY.append(int(racer.text.split(" ")[0]))
        else:
            KVY.append('DNF')
        KVY.append(str(racer.text.split(" ")[9]))
        KVY.append(int(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '7':
        if racer.text.split(" ")[0] != 'DNF':
            RAI.append(int(racer.text.split(" ")[0]))
        else:
            RAI.append('DNF')
        RAI.append(str(racer.text.split(" ")[10]))
        RAI.append(int(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '20':
        if racer.text.split(" ")[0] != 'DNF':
            MAG.append(int(racer.text.split(" ")[0]))
        else:
            MAG.append('DNF')
        MAG.append(str(racer.text.split(" ")[9]))
        MAG.append(int(racer.text.split(" ")[8]))
    # elif racer.text.split(" ")[1] == '8':
    #     if racer.text.split(" ")[0] != 'DNF':
    #         GRO.append(int(racer.text.split(" ")[0]))
    #     else:
    #         GRO.append('DNF')
    #     GRO.append(str(racer.text.split(" ")[9]))
    #     GRO.append(int(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '99':
        if racer.text.split(" ")[0] != 'DNF':
            GIO.append(int(racer.text.split(" ")[0]))
        else:
            GIO.append('DNF')
        GIO.append(str(racer.text.split(" ")[10]))
        GIO.append(int(racer.text.split(" ")[9]))
    elif racer.text.split(" ")[1] == '10':
        if racer.text.split(" ")[0] != 'DNF':
            GAS.append(int(racer.text.split(" ")[0]))
        else:
            GAS.append('DNF')
        GAS.append(str(racer.text.split(" ")[9]))
        GAS.append(int(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '63':
        if racer.text.split(" ")[0] != 'DNF':
            RUS.append(int(racer.text.split(" ")[0]))
        else:
            RUS.append('DNF')
        RUS.append(str(racer.text.split(" ")[9]))
        RUS.append(int(racer.text.split(" ")[8]))
    elif racer.text.split(" ")[1] == '6':
        if racer.text.split(" ")[0] != 'DNF':
            LAT.append(int(racer.text.split(" ")[0]))
        else:
            LAT.append('DNF')
        LAT.append(str(racer.text.split(" ")[8]))
        LAT.append(int(racer.text.split(" ")[7]))
    elif racer.text.split(" ")[1] == '31':
        if racer.text.split(" ")[0] != 'DNF':
            OCO.append(int(racer.text.split(" ")[0]))
        else:
            OCO.append('DNF')
        OCO.append(str(racer.text.split(" ")[11]))
        OCO.append(int(racer.text.split(" ")[10]))
    elif racer.text.split(" ")[1] == '16':
        if racer.text.split(" ")[0] != 'DNF':
            LEC.append(int(racer.text.split(" ")[0]))
        else:
            LEC.append('DNF')
        LEC.append(str(racer.text.split(" ")[8]))
        LEC.append(int(racer.text.split(" ")[7]))
    elif racer.text.split(" ")[1] == '5':
        if racer.text.split(" ")[0] != 'DNF':
            VET.append(int(racer.text.split(" ")[0]))
        else:
            VET.append('DNF')
        VET.append(str(racer.text.split(" ")[8]))
        VET.append(int(racer.text.split(" ")[7]))


file_path = "/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx"
wb = load_workbook(file_path)
s = wb['R15']

colx = 3
rowx = 3


for rowx in range(3, 23):
    posCol = 2
    timeCol = 5
    lapCol = 6
    if s.cell(rowx, colx).value == 'HAM':
        position = HAM[0]
        time = HAM[1]
        lap = HAM[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'BOT':
        position = BOT[0]
        time = BOT[1]
        lap = BOT[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'VET':
        position = VET[0]
        time = VET[1]
        lap = VET[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'LEC':
        position = LEC[0]
        time = LEC[1]
        lap = LEC[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap

    elif s.cell(rowx, colx).value == 'VER':
        position = VER[0]
        time = VER[1]
        lap = VER[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'ALB':
        position = ALB[0]
        time = ALB[1]
        lap = ALB[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'NOR':
        position = NOR[0]
        time = NOR[1]
        lap = NOR[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'SAI':
        position = SAI[0]
        time = SAI[1]
        lap = SAI[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'RIC':
        position = RIC[0]
        time = RIC[1]
        lap = RIC[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'OCO':
        position = OCO[0]
        time = OCO[1]
        lap = OCO[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'GAS':
        position = GAS[0]
        time = GAS[1]
        lap = GAS[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'KVY':
        position = KVY[0]
        time = KVY[1]
        lap = KVY[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'MAG':
        position = MAG[0]
        time = MAG[1]
        lap = MAG[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    # elif s.cell(rowx, colx).value == 'GRO':
    #     position = GRO[0]
    #     time = GRO[1]
    #     lap = GRO[2]
    #     posLoc = s.cell(rowx, posCol)
    #     timeLoc = s.cell(rowx, timeCol)
    #     lapLoc = s.cell(rowx, lapCol)
    #     posLoc.value = position
    #     timeLoc.value = time
    #     lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'RUS':
        position = RUS[0]
        time = RUS[1]
        lap = RUS[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'LAT':
        position = LAT[0]
        time = LAT[1]
        lap = LAT[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'GIO':
        position = GIO[0]
        time = GIO[1]
        lap = GIO[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'RAI':
        position = RAI[0]
        time = RAI[1]
        lap = RAI[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'PER':
        position = PER[0]
        time = PER[1]
        lap = PER[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap
        
    elif s.cell(rowx, colx).value == 'STR':
        position = STR[0]
        time = STR[1]
        lap = STR[2]
        posLoc = s.cell(rowx, posCol)
        timeLoc = s.cell(rowx, timeCol)
        lapLoc = s.cell(rowx, lapCol)
        posLoc.value = position
        timeLoc.value = time
        lapLoc.value = lap


wb.save('/Users/siddhantsingh/Documents/OneDrive/CSProjects/PythonProjects/F1.xlsx')
